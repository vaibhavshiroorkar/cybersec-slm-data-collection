#!/usr/bin/env python3
"""Mark Cleaned?=Yes in sources/Sources (1).xlsx for every Quantum row whose
cleaned output actually exists under clean_data/Quantum/.

Ground-truth marker: independent of the cleaning driver's own bookkeeping. A row
counts as cleaned if its computed output folder under clean_data/ holds at least
one non-empty .jsonl line.
"""
from __future__ import annotations

import os
import re

PROJECT = r"c:/Users/vaibh/Documents/Main/Projects/cybersec-slm-data-pipeline"
XLSX = os.path.join(PROJECT, "sources", "Sources (1).xlsx")
SHEET = "Finalized"
SUBDOMAIN = "Quantum"

os.chdir(PROJECT)
os.environ["CYBERSEC_SLM_DATA_ROOT"] = PROJECT

from cybersec_slm import core                              # noqa: E402
from cybersec_slm.extraction import sources, fetch         # noqa: E402


def norm_header(c) -> str:
    return re.sub(r"[ \-]+", "_", str(c).strip().lower())


def clean_dir_for(descriptor: dict) -> str:
    """Mirror worker._fetch_one's raw folder, mapped into clean_data/."""
    kind = descriptor["kind"]
    domain = descriptor["domain"]
    if kind in ("hf", "kaggle", "url", "github"):
        ref = descriptor["ref"]
        name = ref.split("/")[-1]
        owner = ref.split("/")[0] if "/" in ref and kind in ("hf", "kaggle") else name
        folder = fetch._folder(domain, owner, name, {owner: 1})
    else:
        folder = os.path.join(core.RAW_DATA, domain, descriptor["slug"])
    rel = os.path.relpath(folder, core.RAW_DATA)
    return os.path.join(core.CLEAN_DATA, rel)


def has_records(d: str) -> bool:
    if not os.path.isdir(d):
        return False
    for r, _dirs, files in os.walk(d):
        for fn in files:
            if fn.endswith(".jsonl"):
                p = os.path.join(r, fn)
                try:
                    with open(p, encoding="utf-8", errors="replace") as f:
                        if any(line.strip() for line in f):
                            return True
                except OSError:
                    pass
    return False


def main():
    from openpyxl import load_workbook
    wb = load_workbook(XLSX)
    ws = wb[SHEET]
    header = {norm_header(c.value): c.column for c in ws[1] if c.value is not None}
    sub_c = header.get("sub_domain")
    cleaned_c = header.get("cleaned?")
    if cleaned_c is None:
        raise SystemExit("no 'Cleaned?' column in sheet")

    # build normalized row dicts so we can derive each descriptor
    hdr_by_col = {c.column: norm_header(c.value) for c in ws[1] if c.value is not None}
    marked, cleaned, missing = 0, [], []
    for row in ws.iter_rows(min_row=2):
        if str(row[sub_c - 1].value).strip() != SUBDOMAIN:
            continue
        rd = {hdr_by_col[c.column]: c.value for c in row if c.column in hdr_by_col}
        desc = sources._row_to_descriptor(rd)
        name = sources._val(rd, "name", "source_name", "title", default="?")
        if desc is None:
            missing.append(name)
            continue
        if has_records(clean_dir_for(desc)):
            cleaned.append(name)
            if str(row[cleaned_c - 1].value).strip().lower() not in ("yes", "y"):
                ws.cell(row=row[0].row, column=cleaned_c, value="Yes")
                marked += 1
        else:
            missing.append(name)

    try:
        wb.save(XLSX)
        saved = XLSX
    except PermissionError:
        saved = XLSX.replace(".xlsx", ".cleaned.xlsx")
        wb.save(saved)
    print(f"cleaned (have records): {len(cleaned)}")
    print(f"not cleaned / no records: {len(missing)}")
    print(f"newly marked Cleaned?=Yes: {marked}")
    print(f"saved -> {saved}")
    print("--- not cleaned ---")
    for m in missing:
        print("  ", m)


if __name__ == "__main__":
    main()
