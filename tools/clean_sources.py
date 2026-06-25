#!/usr/bin/env python3
"""End-to-end driver: fetch + clean every source in the spreadsheet, lay the
cleaned output out per sub-domain, and mark Cleaned?=Yes back in the workbook.

For each row in the chosen sheet this:
  1. maps the row to a source descriptor (extraction.sources._row_to_descriptor),
  2. fetches it, cleans it through the full cleaning pipeline, and writes the
     result to  clean_data/<Sub-Domain>/<source>/...  (folder per sub-domain),
  3. deletes the intermediate raw files,
  4. after the pool drains, runs one cross-source dedup pass, then sets
     Cleaned?=Yes for every row whose clean_data/ folder actually holds records
     (ground-truth marking — safe to re-run; it only ever adds Yes).

Generalizes tools/clean_quantum.py (which was Quantum-only) to all sub-domains.

Usage (run from anywhere; paths are pinned below):
    python tools/clean_sources.py                       # all sub-domains
    python tools/clean_sources.py --subdomain "Cloud Security"
    python tools/clean_sources.py --sheet Finalized --workers 6
    python tools/clean_sources.py dryrun                # preview, fetch nothing
    python tools/clean_sources.py mark                  # only re-mark from clean_data/
"""
from __future__ import annotations

import argparse
import importlib.util
import json
import multiprocessing as mp
import os
import re
import subprocess
import sys
import urllib.request
from collections import Counter
from concurrent.futures import ProcessPoolExecutor, as_completed

PROJECT = r"c:/Users/vaibh/Documents/Main/Projects/cybersec-slm-data-pipeline"
XLSX = os.path.join(PROJECT, "sources", "Sources (1).xlsx")
DEFAULT_SHEET = "Finalized"
RESULTS_JSON = os.path.join(PROJECT, "logs", "clean_sources_results.json")

os.makedirs(os.path.join(PROJECT, "logs"), exist_ok=True)
os.chdir(PROJECT)
os.environ["CYBERSEC_SLM_DATA_ROOT"] = PROJECT

from cybersec_slm import core                              # noqa: E402
from cybersec_slm.cleaning import pipeline                 # noqa: E402
from cybersec_slm.extraction import sources, worker        # noqa: E402


# --------------------------------------------------------------- dependencies --
# import-name -> (pip package, what it does). We install ONLY the modules that
# are actually missing — never the whole extra — so an already-working package
# (e.g. fasttext) is never needlessly rebuilt.
_REQUIRED_MODULES = {
    "ftfy": ("ftfy", "encoding repair (sanitize)"),
    "dateutil": ("python-dateutil", "date parsing (sanitize)"),
    "datasketch": ("datasketch", "near-dup MinHash/LSH (dedup)"),
    "langdetect": ("langdetect", "language id fallback (langfilter)"),
    "deep_translator": ("deep-translator", "translate non-English -> English"),
    # best-effort below: build-fragile on Windows / large; pipeline has fallbacks
    "presidio_analyzer": ("presidio-analyzer", "PII detection (pii)"),
    "presidio_anonymizer": ("presidio-anonymizer", "PII redaction (pii)"),
    "fasttext": ("fasttext-wheel", "language id (langfilter)"),
}
# Modules whose absence only downgrades quality — install best-effort, never abort.
_OPTIONAL_MODULES = {"presidio_analyzer", "presidio_anonymizer", "fasttext"}
_SPACY_MODEL = "en_core_web_lg"          # used by presidio for full PII
_FASTTEXT_MODEL = os.path.join(PROJECT, "src", "cybersec_slm", "cleaning", "lid.176.ftz")
_FASTTEXT_URL = ("https://dl.fbaipublicfiles.com/fasttext/supervised-models/"
                 "lid.176.ftz")
_KAGGLE_TOKEN = os.path.join(os.path.expanduser("~"), ".kaggle", "access_token")


def _have(module: str) -> bool:
    try:
        return importlib.util.find_spec(module) is not None
    except (ImportError, ValueError):
        return False


def _pip_install(*args: str) -> None:
    subprocess.check_call([sys.executable, "-m", "pip", "install", *args])


def ensure_deps() -> None:
    """Install everything the cleaning pipeline needs BEFORE any source runs.

    Critical packages (incl. the translator) are a hard requirement — the run
    aborts if they can't be installed, so non-English text is never silently
    dropped. Presidio and the language models are best-effort: missing them only
    downgrades quality (regex PII / langdetect), so they warn instead of abort.
    Idempotent — already-present packages/models are skipped.
    """
    print("[deps] checking cleaning dependencies...", flush=True)
    missing = [m for m in _REQUIRED_MODULES if not _have(m)]
    if not missing:
        print("[deps] all cleaning packages present", flush=True)
    else:
        crit = [m for m in missing if m not in _OPTIONAL_MODULES]
        opt = [m for m in missing if m in _OPTIONAL_MODULES]
        print("[deps] missing packages:", flush=True)
        for m in missing:
            tag = "optional" if m in _OPTIONAL_MODULES else "REQUIRED"
            print(f"         - {m} ({tag}): {_REQUIRED_MODULES[m][1]}", flush=True)

        # Required: must succeed or we abort (so non-English is never dropped).
        if crit:
            pkgs = [_REQUIRED_MODULES[m][0] for m in crit]
            print(f"[deps] installing required: {', '.join(pkgs)}", flush=True)
            try:
                _pip_install(*pkgs)
            except subprocess.CalledProcessError as ex:
                raise SystemExit(f"[deps] FAILED to install required packages: {ex}\n"
                                 "Fix the error above and re-run.") from ex
        # Optional: install one at a time so a build failure (e.g. fasttext needs
        # a C++ compiler on Windows) only downgrades that one stage.
        for m in opt:
            pkg = _REQUIRED_MODULES[m][0]
            print(f"[deps] installing optional: {pkg}", flush=True)
            try:
                _pip_install(pkg)
            except subprocess.CalledProcessError:
                print(f"[deps] WARNING: '{pkg}' install failed (needs a build "
                      f"toolchain?); continuing with the pipeline's fallback.",
                      flush=True)
        importlib.invalidate_caches()

    # spaCy model for presidio (PII falls back to regex without it) — best-effort
    if _have("presidio_analyzer") and not _have(_SPACY_MODEL):
        print(f"[deps] downloading spaCy model {_SPACY_MODEL} (~600MB, one-time)...",
              flush=True)
        try:
            subprocess.check_call([sys.executable, "-m", "spacy", "download", _SPACY_MODEL])
            importlib.invalidate_caches()
        except subprocess.CalledProcessError:
            print("[deps] WARNING: spaCy model download failed; PII uses regex.",
                  flush=True)

    # fastText language-id model (kept out of git) — best-effort (langdetect covers it)
    if not os.path.exists(_FASTTEXT_MODEL):
        print("[deps] downloading fastText lid.176 model...", flush=True)
        os.makedirs(os.path.dirname(_FASTTEXT_MODEL), exist_ok=True)
        try:
            urllib.request.urlretrieve(_FASTTEXT_URL, _FASTTEXT_MODEL)
        except Exception:
            print("[deps] WARNING: fastText model download failed; language id uses "
                  "langdetect/heuristic.", flush=True)

    # Kaggle credentials (kaggle-kind sources fail to fetch without them)
    if not os.path.exists(_KAGGLE_TOKEN) and not os.environ.get("KAGGLE_API_TOKEN"):
        print("[deps] WARNING: no Kaggle token (~/.kaggle/access_token or "
              "KAGGLE_API_TOKEN); kaggle sources will be skipped/failed.", flush=True)

    _report_backends()


def _report_backends() -> None:
    """Print the backend each cleaning stage actually resolved to."""
    from cybersec_slm.cleaning.dedup import Deduper
    from cybersec_slm.cleaning.langfilter import LangFilter
    from cybersec_slm.cleaning.pii import Redactor
    from cybersec_slm.cleaning.translate import Translator
    print(f"[deps] backends -> dedup:{Deduper().backend} pii:{Redactor().engine} "
          f"lang:{LangFilter().backend} translate:{Translator().backend}", flush=True)
    if Translator().backend == "none":
        raise SystemExit("[deps] translator backend is still 'none' after install — "
                         "aborting so non-English text is not silently dropped.")


def norm_header(c) -> str:
    return re.sub(r"[ \-]+", "_", str(c).strip().lower())


# --------------------------------------------------------------- spreadsheet --
def load_rows(sheet: str, subdomain: str | None):
    """Return [{excel_row, name, url, sub_domain, descriptor|None, already}]."""
    from openpyxl import load_workbook
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))   # read_only EmptyCell has no .row
    header = [norm_header(v) if v is not None else "" for v in rows[0]]
    out = []
    for idx, r in enumerate(rows[1:], start=2):   # excel row numbers are 1-based
        if not r or all(v is None for v in r):    # fully blank row
            continue
        rd = {header[i]: (v if v is not None else "")
              for i, v in enumerate(r) if i < len(header) and header[i]}
        sub = str(rd.get("sub_domain", "")).strip()
        if subdomain is not None and sub != subdomain:
            continue
        url = sources._val(rd, "url", "dataset_link", "link", "source_url", default="")
        name = sources._val(rd, "name", "source_name", "title", default=url)
        desc = sources._row_to_descriptor(rd)
        if desc is not None:
            desc = dict(desc)
            desc["_excel_row"] = idx
            desc["_xname"] = name
            desc["_xurl"] = url
        cleaned_flag = str(rd.get("cleaned?", "")).strip().lower()
        out.append({"excel_row": idx, "name": name, "url": url,
                    "sub_domain": sub, "descriptor": desc,
                    "already": cleaned_flag in ("yes", "y", "true", "1")})
    wb.close()
    return out


def clean_dir_for(descriptor: dict) -> str:
    """The clean_data/ folder a descriptor's output lands in (mirrors worker).

    Pure path computation — mirrors fetch._folder's naming (base == owner when a
    single source uses that owner) without creating any directories.
    """
    kind = descriptor["kind"]
    domain = descriptor["domain"]
    if kind in ("hf", "kaggle", "url", "github"):
        ref = descriptor["ref"]
        name = ref.split("/")[-1]
        owner = ref.split("/")[0] if "/" in ref and kind in ("hf", "kaggle") else name
        sub = owner
    else:
        sub = descriptor["slug"]
    return os.path.join(core.CLEAN_DATA, domain, sub)


def has_records(d: str) -> bool:
    if not os.path.isdir(d):
        return False
    for root, _dirs, files in os.walk(d):
        for fn in files:
            if not fn.endswith(".jsonl"):
                continue
            try:
                with open(os.path.join(root, fn), encoding="utf-8", errors="replace") as f:
                    if any(line.strip() for line in f):
                        return True
            except OSError:
                pass
    return False


def mark_spreadsheet(sheet: str, excel_rows: set[int]) -> None:
    """Set Cleaned?=Yes for the given Excel row numbers in `sheet`."""
    if not excel_rows:
        print("[mark] nothing to mark (no source produced records)", flush=True)
        return
    from openpyxl import load_workbook
    wb = load_workbook(XLSX)                  # full load preserves other sheets
    ws = wb[sheet]
    header = {norm_header(c.value): c.column for c in ws[1] if c.value is not None}
    col = header.get("cleaned?")
    if col is None:
        print(f"[mark] WARNING: no 'Cleaned?' column in sheet '{sheet}'; skipping", flush=True)
        return
    n = 0
    for row in ws.iter_rows(min_row=2):
        if row[0].row in excel_rows:
            cur = str(row[col - 1].value).strip().lower()
            if cur not in ("yes", "y"):
                ws.cell(row=row[0].row, column=col, value="Yes")
                n += 1
    try:
        wb.save(XLSX)
        print(f"[mark] set Cleaned?=Yes for {n} rows in {os.path.basename(XLSX)}", flush=True)
    except PermissionError:
        alt = XLSX.replace(".xlsx", ".cleaned.xlsx")
        wb.save(alt)
        print(f"[mark] {XLSX} is locked (open in Excel?); wrote {alt} instead", flush=True)


# ------------------------------------------------------------------ actions ---
def dryrun(sheet: str, subdomain: str | None):
    rows = load_rows(sheet, subdomain)
    print(f"sheet '{sheet}'"
          + (f" sub-domain '{subdomain}'" if subdomain else " (all sub-domains)"))
    print("rows considered :", len(rows))
    print("mappable        :", sum(1 for r in rows if r["descriptor"]))
    print("no-descriptor   :", sum(1 for r in rows if r["descriptor"] is None))
    print("already Yes     :", sum(1 for r in rows if r["already"]))
    print("by sub-domain   :",
          dict(Counter(r["sub_domain"] for r in rows if r["descriptor"])))
    print("by kind         :",
          dict(Counter(r["descriptor"]["kind"] for r in rows if r["descriptor"])))


def mark_only(sheet: str, subdomain: str | None):
    """No fetching — just mark rows whose clean_data/ folder already has records."""
    rows = load_rows(sheet, subdomain)
    hit = {r["excel_row"] for r in rows
           if r["descriptor"] and has_records(clean_dir_for(r["descriptor"]))}
    print(f"[mark] {len(hit)} rows have records under clean_data/", flush=True)
    mark_spreadsheet(sheet, hit)


def run(sheet: str, subdomain: str | None, workers: int | None,
        skip_deps: bool = False):
    if not skip_deps:
        ensure_deps()
    rows = load_rows(sheet, subdomain)
    scope = subdomain or "all sub-domains"
    print(f"[clean] sheet '{sheet}' / {scope}: {len(rows)} rows", flush=True)

    todo = [r for r in rows if r["descriptor"] is not None and not r["already"]]
    skipped = [r for r in rows if r["descriptor"] is None]
    already = [r for r in rows if r["already"] and r["descriptor"] is not None]
    print(f"[clean] to process: {len(todo)}  already-Yes: {len(already)}  "
          f"no-descriptor: {len(skipped)}", flush=True)
    if not todo:
        print("[clean] nothing to do; running mark pass only", flush=True)

    descriptors = [r["descriptor"] for r in todo]
    workers = workers or min(os.cpu_count() or 4, 6)
    results: dict[int, dict] = {}
    done = 0
    ctx = mp.get_context("spawn")
    if descriptors:
        with ProcessPoolExecutor(max_workers=workers, mp_context=ctx) as ex:
            futs = {ex.submit(worker.process_source, d, data_root=PROJECT,
                              clean_data_dir=core.CLEAN_DATA, keep_raw=False): d
                    for d in descriptors}
            for fut in as_completed(futs):
                d = futs[fut]
                er = d["_excel_row"]
                label = d.get("_xname") or d.get("ref") or d.get("slug")
                try:
                    meta = fut.result()
                    rep = meta.get("clean_report_rows", [])
                    out_recs = sum(int(x.get("out", 0)) for x in rep)
                    in_recs = sum(int(x.get("in", 0)) for x in rep)
                    status, err = meta.get("status"), meta.get("error")
                except Exception as ex2:
                    status, out_recs, in_recs, err = "failed", 0, 0, \
                        f"{type(ex2).__name__}: {ex2}"
                results[er] = {"excel_row": er, "name": label, "url": d.get("_xurl"),
                               "sub_domain": d.get("domain"), "kind": d.get("kind"),
                               "status": status, "in": in_recs, "out": out_recs,
                               "error": err}
                done += 1
                print(f"[{done}/{len(todo)}] row{er} {str(d.get('kind')):8} "
                      f"{status:6} in={in_recs} out={out_recs}  {str(label)[:55]}",
                      flush=True)
                with open(RESULTS_JSON, "w", encoding="utf-8") as f:
                    json.dump(results, f, indent=2)

    # one cross-source dedup pass over everything written this run
    try:
        pipeline.final_global_dedup(core.CLEAN_DATA)
    except Exception as ex2:
        print(f"[clean] final dedup skipped: {ex2}", flush=True)

    for r in skipped:
        results[r["excel_row"]] = {"excel_row": r["excel_row"], "name": r["name"],
                                   "url": r["url"], "sub_domain": r["sub_domain"],
                                   "kind": None, "status": "skipped_no_descriptor",
                                   "in": 0, "out": 0,
                                   "error": "could not map row to a source"}
    with open(RESULTS_JSON, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2)

    # ground-truth marking: any selected row whose clean_data/ folder has records
    cleaned_rows = {r["excel_row"] for r in rows
                    if r["descriptor"] and has_records(clean_dir_for(r["descriptor"]))}
    mark_spreadsheet(sheet, cleaned_rows)

    ok = [v for v in results.values() if v["status"] == "ok" and v["out"] > 0]
    zero = [v for v in results.values() if v["status"] == "ok" and v["out"] == 0]
    fail = [v for v in results.values() if v["status"] not in ("ok",)]
    print("\n===== SUMMARY =====", flush=True)
    print(f"cleaned (out>0)      : {len(ok)}", flush=True)
    print(f"ok but 0 records     : {len(zero)}", flush=True)
    print(f"failed/skipped       : {len(fail)}", flush=True)
    print(f"already Yes (skipped): {len(already)}", flush=True)
    print(f"rows marked Cleaned? : {len(cleaned_rows)}", flush=True)
    print(f"results -> {RESULTS_JSON}", flush=True)


def main():
    p = argparse.ArgumentParser(description="Fetch+clean spreadsheet sources, mark Cleaned?=Yes")
    p.add_argument("action", nargs="?", default="run", choices=["run", "dryrun", "mark"])
    p.add_argument("--sheet", default=DEFAULT_SHEET, help="worksheet name (default: Finalized)")
    p.add_argument("--subdomain", default=None, help="limit to one Sub-Domain (default: all)")
    p.add_argument("--workers", type=int, default=None, help="process pool size")
    p.add_argument("--skip-deps", action="store_true",
                   help="skip the pre-run dependency install/check")
    p.add_argument("--deps-only", action="store_true",
                   help="install/verify all dependencies, then exit")
    args = p.parse_args()

    if args.deps_only:
        ensure_deps()
    elif args.action == "dryrun":
        dryrun(args.sheet, args.subdomain)
    elif args.action == "mark":
        mark_only(args.sheet, args.subdomain)
    else:
        run(args.sheet, args.subdomain, args.workers, skip_deps=args.skip_deps)


if __name__ == "__main__":
    main()
