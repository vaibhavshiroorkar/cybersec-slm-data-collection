#!/usr/bin/env python3
"""Fetch+clean every 'Quantum' sub-domain source in sources/Sources (1).xlsx,
then mark Cleaned?=Yes for the rows that produced cleaned records.

Mirrors cybersec_slm.extraction.parallel.run_streaming but filtered to one
sub-domain and with per-row outcome tracking so the spreadsheet can be updated.
"""
from __future__ import annotations

import json
import multiprocessing as mp
import os
import re
import sys
from concurrent.futures import ProcessPoolExecutor, as_completed

PROJECT = r"c:/Users/vaibh/Documents/Main/Projects/cybersec-slm-data-pipeline"
XLSX = os.path.join(PROJECT, "sources", "Sources (1).xlsx")
SHEET = "Finalized"
SUBDOMAIN = "Quantum"
RESULTS_JSON = os.path.join(PROJECT, "logs", "quantum_results.json")

os.makedirs(os.path.join(PROJECT, "logs"), exist_ok=True)
os.chdir(PROJECT)
os.environ["CYBERSEC_SLM_DATA_ROOT"] = PROJECT

from cybersec_slm import core                              # noqa: E402
from cybersec_slm.extraction import sources, worker        # noqa: E402
from cybersec_slm.cleaning import pipeline                 # noqa: E402


def norm_header(c: str) -> str:
    return re.sub(r"[ \-]+", "_", str(c).strip().lower())


def load_quantum_rows():
    """Return list of dicts: {excel_row, name, url, descriptor|None}."""
    from openpyxl import load_workbook
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=False))
    header = [norm_header(c.value) if c.value is not None else "" for c in rows[0]]
    out = []
    for r in rows[1:]:
        rowdict = {header[i]: (c.value if c.value is not None else "")
                   for i, c in enumerate(r) if header[i]}
        sub = str(rowdict.get("sub_domain", "")).strip()
        if sub != SUBDOMAIN:
            continue
        excel_row = r[0].row
        url = sources._val(rowdict, "url", "dataset_link", "link", "source_url", default="")
        name = sources._val(rowdict, "name", "source_name", "title", default=url)
        cleaned_flag = str(rowdict.get("cleaned?", "")).strip().lower()
        desc = sources._row_to_descriptor(rowdict)
        if desc is not None:
            desc = dict(desc)
            desc["_excel_row"] = excel_row
            desc["_xname"] = name
            desc["_xurl"] = url
        out.append({"excel_row": excel_row, "name": name, "url": url,
                    "descriptor": desc, "already": cleaned_flag in ("yes", "y", "true", "1")})
    wb.close()
    return out


def dryrun():
    from collections import Counter
    rows = load_quantum_rows()
    print("total quantum rows:", len(rows))
    print("no-descriptor:", sum(1 for r in rows if r["descriptor"] is None))
    print("already Yes:", sum(1 for r in rows if r["already"]))
    print("kinds:", Counter(r["descriptor"]["kind"] for r in rows if r["descriptor"]))
    print("--- no-descriptor rows ---")
    for r in rows:
        if r["descriptor"] is None:
            print(r["excel_row"], "|", str(r["name"])[:40], "|", str(r["url"])[:60])


def mark_spreadsheet(cleaned_rows: set[int]) -> None:
    """Set Cleaned?=Yes in the Finalized sheet for the given Excel row numbers."""
    if not cleaned_rows:
        print("[quantum] nothing to mark (no source produced records)", flush=True)
        return
    from openpyxl import load_workbook
    wb = load_workbook(XLSX)                       # full load (preserves other sheets/cells)
    ws = wb[SHEET]
    header = {norm_header(c.value): c.column for c in ws[1] if c.value is not None}
    col = header.get("cleaned?")
    if col is None:
        print("[quantum] WARNING: no 'Cleaned?' column found; skipping marking", flush=True)
        return
    n = 0
    for r in sorted(cleaned_rows):
        ws.cell(row=r, column=col, value="Yes")
        n += 1
    try:
        wb.save(XLSX)
        print(f"[quantum] marked Cleaned?=Yes for {n} rows in {os.path.basename(XLSX)}", flush=True)
    except PermissionError:
        alt = XLSX.replace(".xlsx", ".cleaned.xlsx")
        wb.save(alt)
        print(f"[quantum] {XLSX} is locked (open in Excel?); wrote {alt} instead", flush=True)


def main():
    rows = load_quantum_rows()
    print(f"[quantum] {len(rows)} rows under sub-domain '{SUBDOMAIN}'", flush=True)

    todo = [r for r in rows if r["descriptor"] is not None and not r["already"]]
    skipped_build = [r for r in rows if r["descriptor"] is None]
    already = [r for r in rows if r["already"] and r["descriptor"] is not None]
    print(f"[quantum] to process: {len(todo)}  already-Yes: {len(already)}  "
          f"no-descriptor: {len(skipped_build)}", flush=True)

    descriptors = [r["descriptor"] for r in todo]
    workers = min(os.cpu_count() or 4, 6)
    results = {}                       # excel_row -> outcome dict
    ctx = mp.get_context("spawn")
    done = 0
    with ProcessPoolExecutor(max_workers=workers, mp_context=ctx) as ex:
        futs = {ex.submit(worker.process_source, d, data_root=PROJECT,
                          clean_data_dir=core.CLEAN_DATA, keep_raw=False): d
                for d in descriptors}
        for fut in as_completed(futs):
            d = futs[fut]
            er = d["_excel_row"]
            label = d.get("_xname") or d.get("ref") or d.get("slug")
            try:
                meta = fut.result()
                status = meta.get("status")
                rep = meta.get("clean_report_rows", [])
                out_recs = sum(int(x.get("out", 0)) for x in rep)
                in_recs = sum(int(x.get("in", 0)) for x in rep)
                err = meta.get("error")
            except Exception as ex2:
                status, out_recs, in_recs, err, rep = "failed", 0, 0, \
                    f"{type(ex2).__name__}: {ex2}", []
            results[er] = {"excel_row": er, "name": label, "url": d.get("_xurl"),
                           "kind": d.get("kind"), "status": status,
                           "in": in_recs, "out": out_recs, "error": err}
            done += 1
            print(f"[{done}/{len(todo)}] row{er} {d.get('kind'):8} "
                  f"{status:6} in={in_recs} out={out_recs}  {str(label)[:55]}",
                  flush=True)
            with open(RESULTS_JSON, "w", encoding="utf-8") as f:
                json.dump(results, f, indent=2)

    # final cross-source dedup over the freshly written clean_data/
    try:
        pipeline.final_global_dedup(core.CLEAN_DATA)
    except Exception as ex2:
        print(f"[quantum] final dedup skipped: {ex2}", flush=True)

    # add skipped (no descriptor) entries so the report is complete
    for r in skipped_build:
        results[r["excel_row"]] = {"excel_row": r["excel_row"], "name": r["name"],
                                   "url": r["url"], "kind": None,
                                   "status": "skipped_no_descriptor",
                                   "in": 0, "out": 0, "error": "could not map row to a source"}
    with open(RESULTS_JSON, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2)

    ok = [v for v in results.values() if v["status"] == "ok" and v["out"] > 0]
    zero = [v for v in results.values() if v["status"] == "ok" and v["out"] == 0]
    fail = [v for v in results.values() if v["status"] not in ("ok",)]

    mark_spreadsheet({v["excel_row"] for v in ok})

    print("\n===== SUMMARY =====", flush=True)
    print(f"cleaned (out>0): {len(ok)}", flush=True)
    print(f"ok but 0 records: {len(zero)}", flush=True)
    print(f"failed/skipped: {len(fail)}", flush=True)
    print(f"already Yes (untouched): {len(already)}", flush=True)
    print(f"results -> {RESULTS_JSON}", flush=True)


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "dryrun":
        dryrun()
    else:
        main()
